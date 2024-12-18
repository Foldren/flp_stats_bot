import json
from datetime import date, datetime
from io import BytesIO
from aiogram.types import CallbackQuery, BufferedInputFile
from aiogram_dialog import DialogManager, ShowMode
from aiogram_dialog.widgets.kbd import Button
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment
from pytz import timezone
from tortoise.expressions import Q
from config import EXCEL_TEMPLATE_PATH
from models import Bank, PaymentAccount, Transaction


async def on_select_bank(callback: CallbackQuery, button: Button, dialog_manager: DialogManager, selected_bank_id: str):
    bank = await Bank.filter(id=selected_bank_id).first()
    dialog_manager.dialog_data['bank_name'] = bank.name
    dialog_manager.dialog_data['bank_id'] = bank.id
    await dialog_manager.next()


async def on_select_start_date(callback: CallbackQuery, widget, dialog_manager: DialogManager, selected_date: date):
    dialog_manager.dialog_data["start_date"] = selected_date.__str__()
    await dialog_manager.next()


async def on_select_end_date(callback: CallbackQuery, widget, dialog_manager: DialogManager, selected_date: date):
    bank_id = dialog_manager.dialog_data["bank_id"]
    start_date = dialog_manager.dialog_data["start_date"]
    end_date = selected_date.__str__()
    d_start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    d_end_dt = datetime.strptime(end_date, '%Y-%m-%d')
    expression = (Q(p_account__bank_id=bank_id) &
                  (Q(time__range=[d_start_dt, d_end_dt]) | Q(time__startswith=start_date) | Q(time__startswith=end_date)))

    p_accounts = await PaymentAccount.filter(bank_id=bank_id).all()

    if not p_accounts:
        await callback.message.answer("⛔️ Пока нет ниодного расчетного счета по данному банку, подождите подгрузку.")
        await callback.message.delete()
        await dialog_manager.done()
        return

    transactions = await Transaction.filter(expression).select_related("p_account").order_by("time").all()

    if not transactions:
        await callback.message.answer("⛔️ За данный период нет транзакций.")
        await callback.message.delete()
        await dialog_manager.done()
        return

    output = BytesIO()
    wb = load_workbook(filename=EXCEL_TEMPLATE_PATH)
    pa_trxns = {}
    for trxn in transactions:
        try:
            pa_trxns[str(trxn.p_account.number) + " " + trxn.p_account.currency].append(trxn)
        except KeyError:
            pa_trxns[str(trxn.p_account.number) + " " + trxn.p_account.currency] = [trxn]

    for pa_number_cur, transactions in pa_trxns.items():
        # Устанавливаем имя листа
        ws = wb.copy_worksheet(wb["start"])
        ws.title = pa_number_cur

        for i, trxn in enumerate(transactions, start=2):
            ws[f'A{i}'] = trxn.trxn_id if trxn.trxn_id is not None else "-"
            ws[f'B{i}'] = trxn.time.strftime('%Y-%m-%d %H:%S')
            ws[f'C{i}'] = trxn.description
            ws[f'D{i}'] = trxn.amount
            ws[f'C{i}'].alignment = Alignment(wrap_text=True)

    file_name = f"Выписки по банку {dialog_manager.dialog_data["bank_name"]} (с {start_date} по {end_date}).xlsx"

    wb.remove(wb["start"])
    wb.save(output)

    await callback.message.answer(text="✅ Список транзакций сгенерирован!")
    await callback.message.answer_document(document=BufferedInputFile(file=output.getvalue(), filename=file_name))
    await dialog_manager.done()
